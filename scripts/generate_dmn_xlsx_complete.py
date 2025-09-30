#!/usr/bin/env python3
"""
Complete XLSX Generation Script for pyDMNrules Compliance
Generates all 4 DMN tables with explicit borders to avoid the None.style bug
"""

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

def apply_borders(sheet, start_row, end_row, start_col, end_col, double_col=None):
    """
    Apply thin borders to all cells, with optional double border separator

    Args:
        sheet: worksheet
        start_row, end_row: row range
        start_col, end_col: column range
        double_col: column index to apply double right border (input/output separator)
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    double_border = Border(
        left=Side(style='thin'),
        right=Side(style='double'),  # Double border on right
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if double_col and col == double_col:
                cell.border = double_border
            else:
                cell.border = thin_border


def create_weight_classification_dmn(output_file="billing-re/shared/dmn-rules/weight_class.dmn.xlsx"):
    """
    Generate weight_class.dmn.xlsx based on 4_Regeln_Gewichtsklassen.xlsx
    Rules: 6 weight class determinations (20A, 20B, 40A-40D)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Glossary Sheet
    glossary = wb.create_sheet("Glossary")
    glossary['A1'] = "Glossary"
    glossary['A1'].font = Font(bold=True, size=14)

    glossary_headers = ["Variable", "Business Concept", "Attribute"]
    for col, header in enumerate(glossary_headers, 1):
        cell = glossary.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    glossary_data = [
        ["Preisraster", "Pricing", "raster"],
        ["Length", "Dimension", "length"],
        ["GrossWeight", "Weight", "grossWeight"],
        ["WeightClass", "Classification", "weightClass"]
    ]

    for row_idx, row_data in enumerate(glossary_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            glossary.cell(row=row_idx, column=col_idx, value=value)

    apply_borders(glossary, 2, 2 + len(glossary_data), 1, 3)

    # 2. Decision Sheet
    decision = wb.create_sheet("Decision")
    decision['A1'] = "Decision"
    decision['A1'].font = Font(bold=True, size=14)

    decision_headers = ["Decisions", "Execute Decision Tables"]
    for col, header in enumerate(decision_headers, 1):
        cell = decision.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    decision['A3'] = "Determine Weight Class"
    decision['B3'] = "WeightClassification"

    apply_borders(decision, 2, 3, 1, 2)

    # 3. WeightClassification Table Sheet
    table = wb.create_sheet("WeightClassification")

    # Row 1, Col 1: Table name
    table['A1'] = "WeightClassification"
    table['A1'].font = Font(bold=True, size=12)

    # Row 2, Col 1: Hit policy
    table['A2'] = "U"  # UNIQUE hit policy
    table['A2'].font = Font(bold=True, size=12)

    # Row 2, Cols 2+: Variable headers
    variables = ["Preisraster", "Length", "GrossWeight", "WeightClass"]
    for col, var in enumerate(variables, 2):  # Start at column 2 (B)
        cell = table.cell(row=2, column=col, value=var)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Rules data (from roadmap: 20A/20B/40A/40B, plus extra 40C/40D from XLSX)
    # Row 3+, col 1 = rule ID, data starts from column 2
    # Input values quoted, output values UNquoted (pyDMNrules will auto-quote outputs)
    rules = [
        ['"N"', '"20"', '<=20', '20A'],
        ['"N"', '"20"', '>20', '20B'],
        ['"N"', '"40"', '<=10', '40A'],
        ['"N"', '"40"', '[10..20]', '40B'],
        ['"N"', '"40"', '[20..30]', '40C'],
        ['"N"', '"40"', '>30', '40D']
    ]

    for rule_idx, rule_data in enumerate(rules, 3):
        # Add rule ID in column 1 (required by pyDMNrules to identify rules vs validity rows)
        table.cell(row=rule_idx, column=1, value=str(rule_idx - 2))  # Rule IDs: 1, 2, 3...
        # Add rule data starting from column 2
        for col_idx, value in enumerate(rule_data, 2):
            table.cell(row=rule_idx, column=col_idx, value=value)

    # Apply borders with double separator after column 4 (last input before output)
    # Columns: 1=hit policy/empty, 2-4=inputs, 5=output
    apply_borders(table, 1, 2 + len(rules), 1, 5, double_col=4)

    # Auto-width
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    # Save
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"✅ Created: {output_file}")
    return output_file


def create_service_determination_dmn(output_file="billing-re/shared/dmn-rules/service_determination.dmn.xlsx"):
    """
    Generate service_determination.dmn.xlsx based on 3_Regeln_Leistungsermittlung
    COLLECT hit policy (multiple services can apply)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Glossary
    glossary = wb.create_sheet("Glossary")
    glossary['A1'] = "Glossary"
    glossary['A1'].font = Font(bold=True, size=14)

    glossary_headers = ["Variable", "Business Concept", "Attribute"]
    for col, header in enumerate(glossary_headers, 1):
        cell = glossary.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    glossary_data = [
        ["ServiceType", "Service", "type"],
        ["TransportType", "Transport", "transportType"],
        ["LoadingStatus", "Loading", "status"],
        ["DangerousGood", "Cargo", "dangerousGood"],
        ["DepartureStation", "Departure", "station"],
        ["DestinationStation", "Destination", "station"],
        ["CustomsStatus", "Customs", "status"],
        ["CustomsCountry", "Country", "code"],
        ["DateOfService", "DateTime", "serviceDate"],
        ["AdditionalServiceQuantity", "Quantity", "amount"],
        ["AdditionalServiceCode", "ServiceCode", "code"]
    ]

    for row_idx, row_data in enumerate(glossary_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            glossary.cell(row=row_idx, column=col_idx, value=value)

    apply_borders(glossary, 2, 2 + len(glossary_data), 1, 3)

    # 2. Decision
    decision = wb.create_sheet("Decision")
    decision['A1'] = "Decision"
    decision['A1'].font = Font(bold=True, size=14)

    decision_headers = ["Decisions", "Execute Decision Tables"]
    for col, header in enumerate(decision_headers, 1):
        cell = decision.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    decision['A3'] = "Determine Service Codes"
    decision['B3'] = "ServiceDetermination"

    apply_borders(decision, 2, 3, 1, 2)

    # 3. ServiceDetermination Table
    table = wb.create_sheet("ServiceDetermination")

    # Row 1: Table name
    table['A1'] = "ServiceDetermination"
    table['A1'].font = Font(bold=True, size=12)

    # Row 2, Col 1: Hit policy
    table['A2'] = "C"  # COLLECT - multiple services can apply
    table['A2'].font = Font(bold=True, size=12)

    # Row 2, Cols 2+: Variable headers (11 inputs, 1 output)
    variables = [
        "ServiceType",
        "TransportType",
        "LoadingStatus",
        "DangerousGood",
        "DepartureStation",
        "DestinationStation",
        "CustomsStatus",
        "CustomsCountry",
        "DateOfService",
        "DateOfService",
        "AdditionalServiceQuantity",
        "AdditionalServiceCode"
    ]

    for col, var in enumerate(variables, 2):  # Start at column 2
        cell = table.cell(row=2, column=col, value=var)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Rules (from roadmap + XLSX)
    # Using "-" for "don't care" and empty "" for any value
    # Output values (last column) without quotes
    rules = [
        # Rule 1: KV + Dangerous + Date range → 456
        ['-', '"KV"', '-', 'true', '-', '-', '-', '-',
         '>date and time("2025-05-01T00:00:00")',
         '<date and time("2025-08-31T00:00:00")', '-', '456'],

        # Rule 2: Main + Loaded + KV + Dangerous → 456 (alternative)
        ['"MAIN"', '"KV"', '"beladen"', 'true', '-', '-', '-', '-', '-', '-', '-', '456'],

        # Rule 3: Main + Loaded/Empty + KV → 444
        ['"MAIN"', '"KV"', '-', '-', '-', '-', '-', '-', '-', '-', '-', '444'],

        # Rule 4: Main + Any → 111
        ['"MAIN"', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '111'],

        # Rule 5: Trucking + Any → 222
        ['"TRUCKING"', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '222'],

        # Rule 6: Departure station 80155283 → 333
        ['-', '-', '-', '-', '"80155283"', '-', '-', '-', '-', '-', '-', '333'],

        # Rule 7: Destination station 80137943 → 333
        ['-', '-', '-', '-', '-', '"80137943"', '-', '-', '-', '-', '-', '333'],

        # Rule 8: Customs N1 + DE → 555
        ['-', '-', '-', '-', '-', '-', '"N1"', '"DE"', '-', '-', '-', '555'],

        # Rule 9: Additional service with quantity → 789
        ['"ADDITIONAL"', '-', '-', '-', '-', '-', '-', '-', '-', '-', '>0', '789']
    ]

    for rule_idx, rule_data in enumerate(rules, 3):
        # Add rule ID in column 1
        table.cell(row=rule_idx, column=1, value=str(rule_idx - 2))
        # Add rule data starting from column 2
        for col_idx, value in enumerate(rule_data, 2):
            table.cell(row=rule_idx, column=col_idx, value=value)

    # Double border after column 12 (last input before output)
    # Columns: 1=hit policy, 2-12=inputs, 13=output
    apply_borders(table, 1, 2 + len(rules), 1, 13, double_col=12)

    # Auto-width
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 3, 35)

    wb.save(output_file)
    print(f"✅ Created: {output_file}")
    return output_file


def create_trip_type_dmn(output_file="billing-re/shared/dmn-rules/trip_type.dmn.xlsx"):
    """
    Generate trip_type.dmn.xlsx based on 2_Regeln_Fahrttyp
    Simple mapping: trucking code → trip type
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Glossary
    glossary = wb.create_sheet("Glossary")
    glossary['A1'] = "Glossary"
    glossary['A1'].font = Font(bold=True, size=14)

    glossary_headers = ["Variable", "Business Concept", "Attribute"]
    for col, header in enumerate(glossary_headers, 1):
        cell = glossary.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    glossary_data = [
        ["TruckingCode", "Trucking", "code"],
        ["TypeOfTrip", "Trucking", "type"]
    ]

    for row_idx, row_data in enumerate(glossary_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            glossary.cell(row=row_idx, column=col_idx, value=value)

    apply_borders(glossary, 2, 2 + len(glossary_data), 1, 3)

    # 2. Decision
    decision = wb.create_sheet("Decision")
    decision['A1'] = "Decision"
    decision['A1'].font = Font(bold=True, size=14)

    decision_headers = ["Decisions", "Execute Decision Tables"]
    for col, header in enumerate(decision_headers, 1):
        cell = decision.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    decision['A3'] = "Determine Trip Type"
    decision['B3'] = "TripType"

    apply_borders(decision, 2, 3, 1, 2)

    # 3. TripType Table
    table = wb.create_sheet("TripType")

    # Row 1: Table name
    table['A1'] = "TripType"
    table['A1'].font = Font(bold=True, size=12)

    # Row 2, Col 1: Hit policy
    table['A2'] = "U"  # UNIQUE
    table['A2'].font = Font(bold=True, size=12)

    # Row 2, Cols 2+: Variable headers
    variables = ["TruckingCode", "TypeOfTrip"]
    for col, var in enumerate(variables, 2):  # Start at column 2
        cell = table.cell(row=2, column=col, value=var)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Rules from requirement doc
    # Output values (last column) without quotes
    rules = [
        ['"LB"', 'Zustellung'],
        ['"LA"', 'Abholung'],
        ['-', 'Standard']  # Default fallback
    ]

    for rule_idx, rule_data in enumerate(rules, 3):
        # Add rule ID in column 1
        table.cell(row=rule_idx, column=1, value=str(rule_idx - 2))
        # Add rule data starting from column 2
        for col_idx, value in enumerate(rule_data, 2):
            table.cell(row=rule_idx, column=col_idx, value=value)

    # Double border after column 2 (last input before output)
    # Columns: 1=hit policy, 2=input, 3=output
    apply_borders(table, 1, 2 + len(rules), 1, 3, double_col=2)

    # Auto-width
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_file)
    print(f"✅ Created: {output_file}")
    return output_file


def create_tax_calculation_dmn(output_file="billing-re/shared/dmn-rules/tax_calculation.dmn.xlsx"):
    """
    Generate tax_calculation.dmn.xlsx based on 3_1_Regeln_Steuerberechnung
    Tax determination for Export/Import/Domestic scenarios
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Glossary
    glossary = wb.create_sheet("Glossary")
    glossary['A1'] = "Glossary"
    glossary['A1'].font = Font(bold=True, size=14)

    glossary_headers = ["Variable", "Business Concept", "Attribute"]
    for col, header in enumerate(glossary_headers, 1):
        cell = glossary.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    glossary_data = [
        ["TransportDirection", "Transport", "direction"],
        ["CustomsStatus", "Customs", "status"],
        ["DepartureCountry", "DepartureLocation", "country"],
        ["DestinationCountry", "DestinationLocation", "country"],
        ["SupplySubjectToVAT", "VAT", "subjectToVAT"],
        ["VATRate", "VATCalculation", "rate"],
        ["TaxCase", "TaxRule", "case"],
        ["TaxNote", "TaxInfo", "note"]
    ]

    for row_idx, row_data in enumerate(glossary_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            glossary.cell(row=row_idx, column=col_idx, value=value)

    apply_borders(glossary, 2, 2 + len(glossary_data), 1, 3)

    # 2. Decision
    decision = wb.create_sheet("Decision")
    decision['A1'] = "Decision"
    decision['A1'].font = Font(bold=True, size=14)

    decision_headers = ["Decisions", "Execute Decision Tables"]
    for col, header in enumerate(decision_headers, 1):
        cell = decision.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    decision['A3'] = "Determine Tax"
    decision['B3'] = "TaxCalculation"

    apply_borders(decision, 2, 3, 1, 2)

    # 3. TaxCalculation Table
    table = wb.create_sheet("TaxCalculation")

    # Row 1: Table name
    table['A1'] = "TaxCalculation"
    table['A1'].font = Font(bold=True, size=12)

    # Row 2, Col 1: Hit policy
    table['A2'] = "U"  # UNIQUE
    table['A2'].font = Font(bold=True, size=12)

    # Row 2, Cols 2+: Variable headers
    variables = [
        "TransportDirection",
        "CustomsStatus",
        "DepartureCountry",
        "DestinationCountry",
        "SupplySubjectToVAT",
        "VATRate",
        "TaxCase",
        "TaxNote"
    ]

    for col, var in enumerate(variables, 2):  # Start at column 2
        cell = table.cell(row=2, column=col, value=var)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Rules from roadmap (Export 0%, Import reverse, Domestic 19%)
    # Output values (last 4 columns) without quotes
    rules = [
        # Export: DE→US, T1-NCTS → 0% VAT
        ['"Export"', '"T1-NCTS"', '"DE"', '-', 'nein', 0, '§4 Nr. 3a UStG', 'A1'],

        # Import: US→DE, reverse charge
        ['"Import"', '-', '-', '"DE"', 'ja', 0, 'Reverse Charge', 'A2'],

        # Domestic: DE→DE → 19% VAT
        ['"Domestic"', '-', '"DE"', '"DE"', 'ja', 19, 'Standard', 'A3'],

        # Default fallback
        ['-', '-', '-', '-', 'ja', 19, 'Standard', 'Default']
    ]

    for rule_idx, rule_data in enumerate(rules, 3):
        # Add rule ID in column 1
        table.cell(row=rule_idx, column=1, value=str(rule_idx - 2))
        # Add rule data starting from column 2
        for col_idx, value in enumerate(rule_data, 2):
            table.cell(row=rule_idx, column=col_idx, value=value)

    # Double border after column 5 (last input before outputs)
    # Columns: 1=hit policy, 2-5=inputs, 6-9=outputs
    apply_borders(table, 1, 2 + len(rules), 1, 9, double_col=5)

    # Auto-width
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_file)
    print(f"✅ Created: {output_file}")
    return output_file


if __name__ == "__main__":
    print("🚀 Generating pyDMNrules-compliant XLSX files...\n")

    # Create all 4 DMN tables
    create_weight_classification_dmn()
    create_service_determination_dmn()
    create_trip_type_dmn()
    create_tax_calculation_dmn()

    print("\n✅ All DMN XLSX files generated successfully!")
    print("\nKey features:")
    print("  • Explicit thin borders on all cells")
    print("  • Double borders separating inputs from outputs")
    print("  • FEEL syntax for dates, strings, and ranges")
    print("  • Glossary and Decision sheets for metadata")
    print("  • Hit policies: UNIQUE (single result), COLLECT (multiple)")
    print("\nReady to load with: dmn.load('billing-re/shared/dmn-rules/weight_class.dmn.xlsx')")