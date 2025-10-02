#!/usr/bin/env python3
"""
XLSX Tax Calculation Processor
Reads tax rules from shared/rules/3_1_Regeln_Steuerberechnung.xlsx
Implements advanced tax determination based on transport conditions
"""

import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, List
import logging

logger = logging.getLogger(__name__)


class XLSXTaxProcessor:
    """Process tax calculation rules from XLSX files"""

    def __init__(self, rules_dir: Path = None):
        """
        Initialize tax processor with rules directory

        Args:
            rules_dir: Path to directory containing DMN XLSX rules (optional)
        """
        if rules_dir is None:
            # Calculate project root from current file location
            current_file = Path(__file__).resolve()
            billing_service_dir = current_file.parent  # services/billing
            project_root = billing_service_dir.parent.parent  # services -> project_root
            rules_dir = project_root / "shared" / "rules"

        self.rules_dir = Path(rules_dir)
        self.tax_calculation_file = self.rules_dir / "3_1_Regeln_Steuerberechnung.xlsx"

        logger.info(f"XLSXTaxProcessor initialized with rules_dir: {self.rules_dir}")
        logger.info(f"Tax calculation file: {self.tax_calculation_file}")

        if not self.tax_calculation_file.exists():
            logger.warning(f"Tax calculation file not found at {self.tax_calculation_file}")

    def evaluate_tax_calculation(self, order_context: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Evaluate tax calculation based on order context

        XLSX Structure (3_1_Regeln_Steuerberechnung.xlsx):
        - Column 1: Hauptleistung (Main service type)
        - Column 2: Ladezustand (Loading status)
        - Column 3: Versandort (Departure location type: Inland/Ausland)
        - Column 4: Empfangsort (Destination location type: Inland/Ausland)
        - Column 5: USt-ID (VAT ID presence)
        - Column 6: USt-Land (VAT country)
        - Column 7: Transportrichtung (Transport direction: Export/Import/Domestic)
        - Column 8: Zoll-Verfahren (Customs procedure)
        - Column 10: Umsatzsteuer setzen (Apply VAT: ja/nein)
        - Column 11: Steuerfall setzen (Tax case: § reference)
        - Column 12: Hinweis 1 darstellen (Display notice 1: ja/nein)
        - Column 13: SAP USt-Kennzeichen setzen (SAP VAT indicator)
        - Column 14: Angabe Zentrale Meldung (Central notification)

        Args:
            order_context: Dictionary with transport and location data
                Required keys:
                - service_type: Type of main service (e.g., 'Transport')
                - loading_status: Loading status (e.g., 'beladen', 'leer')
                - departure_location_type: 'Inland' or 'Ausland'
                - destination_location_type: 'Inland' or 'Ausland'
                - vat_id: VAT ID presence ('DE', 'keine', etc.)
                - vat_country: Country code or 'nicht relevant'
                - transport_direction: 'Export', 'Import', or 'Domestic'
                - customs_procedure: Procedure type (e.g., 'T1-NCTS', 'GW', etc.)

        Returns:
            Dictionary with tax calculation result or None if no match:
            {
                'apply_vat': bool,
                'tax_case': str,
                'tax_rate': float,
                'display_notice': bool,
                'sap_vat_indicator': str,
                'central_notification': bool,
                'rule_matched': str
            }
        """
        if not self.tax_calculation_file.exists():
            logger.error(f"Tax calculation file not found: {self.tax_calculation_file}")
            return None

        try:
            wb = openpyxl.load_workbook(self.tax_calculation_file, data_only=True)
            ws = wb.active

            # Extract order context values
            service_type = order_context.get('service_type', '').lower()
            loading_status = order_context.get('loading_status', '').lower()
            departure_location = order_context.get('departure_location_type', '').lower()
            destination_location = order_context.get('destination_location_type', '').lower()
            vat_id = order_context.get('vat_id', '').upper()
            vat_country = order_context.get('vat_country', '').upper()
            transport_direction = order_context.get('transport_direction', '').lower()
            customs_procedure = order_context.get('customs_procedure', '').upper()

            logger.info(f"Evaluating tax rules for order: {order_context}")
            logger.info(f"Tax XLSX file: {self.tax_calculation_file}, exists: {self.tax_calculation_file.exists()}")

            # Iterate through rules (starting from row 3, row 2 is header)
            for row_idx in range(3, ws.max_row + 1):
                # Skip placeholder rows
                rule_service = self._get_cell_value(ws, row_idx, 1)
                if rule_service == '…':
                    continue

                # Extract rule conditions
                rule_service = self._normalize_value(rule_service)
                rule_loading = self._normalize_value(self._get_cell_value(ws, row_idx, 2))
                rule_departure = self._normalize_value(self._get_cell_value(ws, row_idx, 3))
                rule_destination = self._normalize_value(self._get_cell_value(ws, row_idx, 4))
                rule_vat_id = self._normalize_value(self._get_cell_value(ws, row_idx, 5))
                rule_vat_country = self._normalize_value(self._get_cell_value(ws, row_idx, 6))
                rule_direction = self._normalize_value(self._get_cell_value(ws, row_idx, 7))
                rule_customs = self._normalize_value(self._get_cell_value(ws, row_idx, 8))

                # Check if all conditions match (None/empty = wildcard)
                matches = []

                # Service type match
                if rule_service and 'transport' in service_type:
                    matches.append('transport' in rule_service.lower())
                elif rule_service:
                    matches.append(False)
                else:
                    matches.append(True)  # Wildcard

                # Loading status match
                if rule_loading:
                    matches.append(rule_loading.lower() in loading_status)
                else:
                    matches.append(True)

                # Departure location match
                if rule_departure:
                    matches.append(rule_departure.lower() in departure_location)
                else:
                    matches.append(True)

                # Destination location match
                # SPECIAL CASE: For Export with "Inland" destination in XLSX, treat as wildcard
                # because XLSX data structure doesn't match physical transport direction
                if rule_destination:
                    if rule_direction and 'export' in rule_direction.lower() and 'inland' in rule_destination.lower():
                        matches.append(True)  # Treat as wildcard for Export
                    else:
                        matches.append(rule_destination.lower() in destination_location)
                else:
                    matches.append(True)

                # VAT ID match
                if rule_vat_id:
                    if 'keine' in rule_vat_id.lower():
                        matches.append('keine' in vat_id.lower() or not vat_id)
                    elif 'relevant' in rule_vat_id.lower():
                        matches.append(True)  # "nicht relevant" = wildcard
                    else:
                        matches.append(rule_vat_id.upper() == vat_id.upper())
                else:
                    matches.append(True)

                # VAT country match
                if rule_vat_country:
                    if 'relevant' in rule_vat_country.lower():
                        matches.append(True)  # "nicht relevant" = wildcard
                    else:
                        # If VAT ID is "keine", treat vat_country as wildcard for Export
                        if vat_id and 'keine' in vat_id.lower() and rule_direction and 'export' in rule_direction.lower():
                            matches.append(True)  # No VAT ID for Export = match any country
                        else:
                            matches.append(rule_vat_country.upper() == vat_country.upper())
                else:
                    matches.append(True)

                # Transport direction match
                if rule_direction:
                    matches.append(rule_direction.lower() in transport_direction.lower())
                else:
                    matches.append(True)

                # Customs procedure match
                if rule_customs:
                    # If order has no customs procedure, treat as wildcard match for Export
                    if not customs_procedure and rule_direction and 'export' in rule_direction.lower():
                        matches.append(True)  # No customs procedure provided = match any Export rule
                    else:
                        # Handle comma-separated list
                        customs_list = [c.strip().upper() for c in rule_customs.split(',')]
                        matches.append(any(c in customs_procedure.upper() for c in customs_list))
                else:
                    matches.append(True)

                # Log matching progress for debugging
                logger.debug(f"Row {row_idx} matching: {matches} - Service:{rule_service}, Loading:{rule_loading}, Dep:{rule_departure}, Dest:{rule_destination}, VAT-ID:{rule_vat_id}, Direction:{rule_direction}, Customs:{rule_customs}")

                # If all conditions match, extract actions
                if all(matches):
                    logger.info(f"Tax rule matched at row {row_idx}")

                    # Extract action columns
                    apply_vat_str = self._get_cell_value(ws, row_idx, 10)
                    tax_case_str = self._get_cell_value(ws, row_idx, 11)
                    display_notice_str = self._get_cell_value(ws, row_idx, 12)
                    sap_indicator = self._get_cell_value(ws, row_idx, 13)
                    central_notification_str = self._get_cell_value(ws, row_idx, 14)

                    # Parse VAT application
                    apply_vat = apply_vat_str and apply_vat_str.lower() in ['ja', 'yes', 'true']

                    # Determine tax rate based on tax case
                    tax_rate = 0.0
                    if apply_vat:
                        # German standard VAT rate
                        tax_rate = 0.19
                    elif tax_case_str and 'reverse' in tax_case_str.lower():
                        tax_rate = 0.0  # Reverse charge

                    return {
                        'apply_vat': apply_vat,
                        'tax_case': tax_case_str or 'No VAT',
                        'tax_rate': tax_rate,
                        'display_notice': display_notice_str and display_notice_str.lower() in ['ja', 'yes', 'true'],
                        'sap_vat_indicator': sap_indicator or '',
                        'central_notification': central_notification_str and central_notification_str.lower() in ['ja', 'yes', 'true'],
                        'rule_matched': f'Row {row_idx}'
                    }

            logger.warning(f"No tax rule matched for order context: {order_context}")
            return None

        except Exception as e:
            logger.error(f"Error evaluating tax calculation: {e}", exc_info=True)
            return None

    def _get_cell_value(self, worksheet, row: int, col: int) -> str:
        """Safely get cell value as string"""
        cell_value = worksheet.cell(row, col).value
        return str(cell_value).strip() if cell_value is not None else ''

    def _normalize_value(self, value: str) -> str:
        """Normalize value by stripping and handling None"""
        if not value or value.lower() in ['none', 'null', '']:
            return ''
        return value.strip()

    def calculate_tax_for_transport(self,
                                   transport_direction: str,
                                   departure_country: str,
                                   destination_country: str,
                                   vat_id: Optional[str] = None,
                                   customs_procedure: Optional[str] = None,
                                   loading_status: str = 'beladen') -> Dict[str, Any]:
        """
        Convenience method for tax calculation based on transport parameters

        Args:
            transport_direction: 'Export', 'Import', or 'Domestic'
            departure_country: ISO country code
            destination_country: ISO country code
            vat_id: VAT ID if available
            customs_procedure: Customs procedure code
            loading_status: Container loading status

        Returns:
            Dictionary with tax calculation result
        """
        # Determine location types
        departure_location = 'Inland' if departure_country.upper() == 'DE' else 'Ausland'
        destination_location = 'Inland' if destination_country.upper() == 'DE' else 'Ausland'

        # Prepare order context
        order_context = {
            'service_type': 'Transport',
            'loading_status': loading_status,
            'departure_location_type': departure_location,
            'destination_location_type': destination_location,
            'vat_id': vat_id or 'keine',
            'vat_country': destination_country.upper(),  # Always use destination country
            'transport_direction': transport_direction,
            'customs_procedure': customs_procedure or ''
        }

        result = self.evaluate_tax_calculation(order_context)

        if result:
            return result

        # Fallback to simple rules
        logger.warning(f"Using fallback tax calculation for {transport_direction}")
        if transport_direction.lower() == 'export':
            return {
                'apply_vat': False,
                'tax_case': '§ 4 Nr. 3a UStG',
                'tax_rate': 0.0,
                'display_notice': False,
                'sap_vat_indicator': 'A1',
                'central_notification': False,
                'rule_matched': 'Fallback Export Rule'
            }
        elif transport_direction.lower() == 'import':
            return {
                'apply_vat': False,
                'tax_case': 'Reverse Charge',
                'tax_rate': 0.0,
                'display_notice': False,
                'sap_vat_indicator': 'RC',
                'central_notification': False,
                'rule_matched': 'Fallback Import Rule'
            }
        else:  # Domestic
            return {
                'apply_vat': True,
                'tax_case': 'Standard VAT',
                'tax_rate': 0.19,
                'display_notice': False,
                'sap_vat_indicator': 'B1',
                'central_notification': False,
                'rule_matched': 'Fallback Domestic Rule'
            }
