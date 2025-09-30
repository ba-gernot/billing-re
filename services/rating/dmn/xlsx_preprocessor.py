"""
XLSX Preprocessor for pyDMNrules
Fixes output cell values to be FEEL-compliant before pyDMNrules loads them

Problem: pyDMNrules treats output cells as FEEL expressions
Solution: Pre-process XLSX to wrap simple string outputs in quotes
"""

import openpyxl
from pathlib import Path
from typing import Optional
import logging
import tempfile
import shutil

logger = logging.getLogger(__name__)


class XLSXPreprocessor:
    """
    Preprocesses DMN XLSX files to make them compatible with pyDMNrules

    Main fix: Wraps unquoted string output values in FEEL quotes
    Example: '20A' → '"20A"'
    """

    def __init__(self):
        self._temp_files = {}

    def preprocess_for_pydmnrules(self, xlsx_path: Path) -> Optional[Path]:
        """
        Preprocess an XLSX file to fix output columns for pyDMNrules

        Args:
            xlsx_path: Path to original XLSX file

        Returns:
            Path to preprocessed XLSX file (in temp directory)
            Returns None if preprocessing fails
        """
        try:
            logger.debug(f"[PREPROC] Preprocessing {xlsx_path.name}")

            # Load workbook
            wb = openpyxl.load_workbook(xlsx_path)

            # Find decision table sheets (skip Glossary and Decision)
            decision_tables = [
                sheet for sheet in wb.sheetnames
                if sheet not in ['Glossary', 'Decision']
            ]

            logger.debug(f"[PREPROC] Found decision tables: {decision_tables}")

            # Process each table
            modified = False
            for table_name in decision_tables:
                if self._fix_output_columns(wb[table_name]):
                    modified = True

            if not modified:
                logger.debug(f"[PREPROC] No changes needed for {xlsx_path.name}")
                return None

            # Save to temporary file
            temp_dir = Path(tempfile.gettempdir()) / "dmn_preprocessed"
            temp_dir.mkdir(exist_ok=True)

            temp_path = temp_dir / xlsx_path.name
            wb.save(str(temp_path))

            logger.info(f"[PREPROC] Preprocessed {xlsx_path.name} → {temp_path}")
            self._temp_files[str(xlsx_path)] = temp_path

            return temp_path

        except Exception as e:
            logger.error(f"[PREPROC] Failed to preprocess {xlsx_path}: {e}")
            return None

    def _fix_output_columns(self, sheet) -> bool:
        """
        Fix output column cells in a decision table

        Identifies output columns and quotes simple string values

        Returns:
            True if any cells were modified
        """
        try:
            # Check if this is a decision table (has hit policy in A2)
            hit_policy = sheet['A2'].value
            if not hit_policy or str(hit_policy).strip() not in ['U', 'A', 'P', 'F', 'R', 'O', 'C', 'C+', 'C<', 'C>', 'C#']:
                logger.debug(f"[PREPROC] Sheet {sheet.title} is not a decision table")
                return False

            logger.debug(f"[PREPROC] Processing table: {sheet.title}, hit policy: {hit_policy}")

            # Parse headers to find input/output columns
            # Headers are in row 2, starting from column 2
            max_col = sheet.max_column
            max_row = sheet.max_row

            # Find where inputs end and outputs begin
            # Look for double border or use glossary to determine
            # For now, use simple heuristic: last 1-2 columns are outputs

            # Better approach: Read all headers and identify output columns
            # Outputs typically have simple names without complex expressions
            headers = []
            for col in range(2, max_col + 1):
                header_cell = sheet.cell(2, col)
                headers.append({
                    'col': col,
                    'name': header_cell.value,
                    'border': header_cell.border
                })

            logger.debug(f"[PREPROC] Headers: {[h['name'] for h in headers]}")

            # Identify output columns (heuristic: last column(s) without comparison operators in data)
            output_cols = self._identify_output_columns(sheet, headers)

            if not output_cols:
                logger.debug(f"[PREPROC] No output columns identified")
                return False

            logger.debug(f"[PREPROC] Output columns: {output_cols}")

            # Fix output cell values
            modified = False
            for row in range(3, max_row + 1):
                for col_info in output_cols:
                    col = col_info['col']
                    cell = sheet.cell(row, col)

                    if self._should_quote_cell(cell):
                        original = cell.value
                        # Wrap in FEEL quotes
                        cell.value = f'"{original}"'
                        logger.debug(f"[PREPROC] Fixed {sheet.title}!{cell.coordinate}: '{original}' → '{cell.value}'")
                        modified = True

            return modified

        except Exception as e:
            logger.error(f"[PREPROC] Error fixing output columns in {sheet.title}: {e}")
            return False

    def _identify_output_columns(self, sheet, headers: list) -> list:
        """
        Identify which columns are outputs vs inputs

        Heuristic: Output columns typically have:
        - Simple string values (e.g., "20A", "Zustellung")
        - No comparison operators (<, >, =, -, [])
        """
        output_cols = []

        # Check data in first few rows to identify patterns
        for col_info in headers:
            col = col_info['col']
            name = col_info['name']

            # Sample first 3 data rows
            sample_values = []
            for row in range(3, min(6, sheet.max_row + 1)):
                cell = sheet.cell(row, col)
                if cell.value:
                    sample_values.append(str(cell.value))

            if not sample_values:
                continue

            # Check if values look like outputs (no operators)
            is_output = all(
                not any(op in val for op in ['<', '>', '=', '-', '[', ']', '(', ')', ','])
                for val in sample_values
            )

            if is_output:
                # Additionally check: outputs are typically last columns
                # and have simpler patterns
                is_simple_string = all(
                    isinstance(sheet.cell(3 + i, col).value, str) and
                    len(sheet.cell(3 + i, col).value) < 30 and
                    not sheet.cell(3 + i, col).value.startswith('"')
                    for i in range(min(3, len(sample_values)))
                    if sheet.cell(3 + i, col).value
                )

                if is_simple_string:
                    output_cols.append(col_info)
                    logger.debug(f"[PREPROC] Identified output column: {name} (col {col})")

        return output_cols

    def _should_quote_cell(self, cell) -> bool:
        """
        Determine if a cell value should be wrapped in quotes

        Criteria:
        - Is a string
        - Not already quoted
        - Not empty
        - Not a FEEL expression (no operators, brackets, etc.)
        """
        if cell.value is None:
            return False

        if not isinstance(cell.value, str):
            return False

        value = cell.value.strip()

        if not value:
            return False

        # Already quoted
        if value.startswith('"') and value.endswith('"'):
            return False

        # Is a FEEL expression (has operators or special syntax)
        if any(char in value for char in ['<', '>', '=', '[', ']', '(', ')', ',']):
            return False

        # Should quote: simple string like "20A", "Zustellung", "Standard"
        return True

    def cleanup(self):
        """Clean up temporary preprocessed files"""
        for temp_path in self._temp_files.values():
            try:
                if temp_path.exists():
                    temp_path.unlink()
                    logger.debug(f"[PREPROC] Cleaned up {temp_path}")
            except Exception as e:
                logger.warning(f"[PREPROC] Failed to cleanup {temp_path}: {e}")

        self._temp_files.clear()


# Global instance
_preprocessor = XLSXPreprocessor()


def get_preprocessor() -> XLSXPreprocessor:
    """Get the global XLSX preprocessor instance"""
    return _preprocessor