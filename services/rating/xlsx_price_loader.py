#!/usr/bin/env python3
"""
XLSX Price Loader - Dynamic price loading from Excel files
Loads prices from XLSX files in shared/price-tables/ directory
Similar to xlsx_dmn_processor.py but for pricing data
"""

import openpyxl
from typing import Dict, Any, Optional, List
from pathlib import Path
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class XLSXPriceLoader:
    """
    Dynamic XLSX price loader that reads pricing tables from Excel files
    Supports automatic file modification detection and caching
    """

    def __init__(self, prices_dir: Path):
        self.prices_dir = Path(prices_dir)
        self._price_cache: Dict[str, Dict] = {}
        self._file_mtimes: Dict[str, float] = {}  # Track file modification times

    def load_price_file(self, file_name: str, force_reload: bool = False) -> Optional[Dict]:
        """Load and parse a price XLSX file with automatic modification detection"""

        file_path = self.prices_dir / file_name

        if not file_path.exists():
            logger.warning(f"Price file not found: {file_path}")
            return None

        # Check if file was modified since last load
        try:
            current_mtime = file_path.stat().st_mtime
        except Exception as e:
            logger.error(f"Failed to get file modification time for {file_name}: {e}")
            current_mtime = 0

        # Use cache if file hasn't been modified and not forcing reload
        if not force_reload and file_name in self._price_cache:
            cached_mtime = self._file_mtimes.get(file_name, 0)
            if current_mtime <= cached_mtime:
                logger.debug(f"Using cached prices for {file_name}")
                return self._price_cache[file_name]
            else:
                logger.info(f"File {file_name} modified (cached: {cached_mtime}, current: {current_mtime}), reloading")

        # Load file and cache with modification time
        try:
            wb = openpyxl.load_workbook(file_path)
            price_data = self._parse_workbook(wb, file_name)
            self._price_cache[file_name] = price_data
            self._file_mtimes[file_name] = current_mtime
            logger.info(f"Loaded prices from {file_name} (mtime: {current_mtime})")
            return price_data

        except Exception as e:
            logger.error(f"Failed to load price file {file_name}: {e}")
            return None

    def _parse_workbook(self, wb: openpyxl.Workbook, file_name: str) -> Dict:
        """Parse a workbook and extract price data"""

        price_data = {
            'file_name': file_name,
            'sheets': {},
            'prices': []
        }

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = self._parse_sheet(sheet, sheet_name)
            price_data['sheets'][sheet_name] = sheet_data

            # Extract prices from sheet
            if sheet_data['headers'] and sheet_data['rows']:
                prices = self._extract_prices_from_sheet(sheet_data, sheet_name)
                price_data['prices'].extend(prices)

        return price_data

    def _parse_sheet(self, sheet, sheet_name: str) -> Dict:
        """Parse a worksheet and extract structured data"""

        sheet_data = {
            'name': sheet_name,
            'headers': [],
            'rows': [],
            'max_row': sheet.max_row,
            'max_column': sheet.max_column
        }

        if sheet.max_row < 1:
            return sheet_data

        # Get headers from first row
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, col).value
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Col_{col}")

        sheet_data['headers'] = headers

        # Get data rows
        for row in range(2, sheet.max_row + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row, col).value
                if cell_value is not None:
                    row_data.append(cell_value)
                else:
                    row_data.append(None)

            # Skip empty rows
            if any(cell is not None for cell in row_data):
                sheet_data['rows'].append(row_data)

        return sheet_data

    def _extract_prices_from_sheet(self, sheet_data: Dict, sheet_name: str) -> List[Dict]:
        """Extract price entries from sheet data"""

        prices = []
        headers = sheet_data['headers']
        rows = sheet_data['rows']

        for i, row in enumerate(rows):
            price_entry = {
                'sheet': sheet_name,
                'row_id': i + 2,  # +2 because row 1 is header, 0-indexed
                'data': {}
            }

            # Map row data to headers
            for j, header in enumerate(headers):
                if j < len(row):
                    value = row[j]
                    if value is not None:
                        price_entry['data'][header] = value

            if price_entry['data']:
                prices.append(price_entry)

        return prices

    def get_main_service_price_advanced(self, customer_code: str, customer_group: str, offer_number: str,
                                        departure_country: str, departure_station: str, tariff_point_dep: str,
                                        destination_country: str, destination_station: str, tariff_point_dest: str,
                                        direction: str, loading_status: str, transport_form: str,
                                        container_length: str, weight_class: str, service_date: str,
                                        **kwargs) -> Optional[Dict[str, Any]]:
        """
        Get main service price with advanced 19-column matching and specificity ranking

        Args:
            All the matching criteria from the pricing XLSX

        Returns:
            Dictionary with price and matching details or None if not found
        """

        # Load from shared folder - fix path resolution
        # self.prices_dir is usually services/rating/price-tables
        # We need to go up to project root, then to shared
        rating_service_dir = self.prices_dir.parent if 'price-tables' in str(self.prices_dir) else self.prices_dir
        project_root = rating_service_dir.parent.parent  # services/rating -> services -> project_root
        shared_path = project_root / "shared" / "rules" / "6_Preistabelle_Hauptleistungen_Einzelpreise.xlsx"

        if not shared_path.exists():
            logger.warning(f"Main service price table not found at {shared_path}")
            return None

        try:
            import openpyxl
            wb = openpyxl.load_workbook(shared_path)
            ws = wb.active

            # Expected columns (19 total):
            # 0: Angebotsnummer, 1: Kundengruppe, 2: Kundennummer,
            # 3: Land Bahnstelle Versand, 4: Bahnstellennummer Versand, 5: Tarifpunkt Versand,
            # 6: Land Bahnstelle Empfang, 7: Bahnstellennummer Empfang, 8: Tarifpunkt Empfang,
            # 9: Richtung, 10: Ladezustand, 11: Verkehrsform,
            # 12: Preisraster, 13: Container Länge, 14: Gewichtsklasse,
            # 15: gültig von, 16: gültig bis, 17: Preis, 18: Preiseinheit

            matches = []

            for row_idx in range(2, ws.max_row + 1):
                row = []
                for col in range(1, ws.max_column + 1):
                    row.append(ws.cell(row_idx, col).value)

                if not any(row):
                    continue

                # Calculate specificity score
                specificity = 0

                # Customer number match (most specific) = +1000
                if row[2] and str(row[2]) == str(customer_code):
                    specificity += 1000
                # Customer group match = +100
                elif row[1] and str(row[1]) == str(customer_group):
                    specificity += 100
                # Offer number match = +50
                elif row[0] and str(row[0]) == str(offer_number):
                    specificity += 50

                # Station matches = +10 each
                if row[4] and str(row[4]) == str(departure_station):
                    specificity += 10
                if row[7] and str(row[7]) == str(destination_station):
                    specificity += 10

                # Tariff point matches = +5 each
                if row[5] and str(row[5]) == str(tariff_point_dep):
                    specificity += 5
                if row[8] and str(row[8]) == str(tariff_point_dest):
                    specificity += 5

                # Required matches (must match, no specificity bonus)
                if row[9] and str(row[9]) != direction:
                    continue  # No match
                if row[14] and str(row[14]) != weight_class:
                    continue  # No match
                if row[13] and str(row[13]) != str(container_length):
                    continue  # No match

                # Optional matches (add small specificity if they match)
                if row[10] and str(row[10]) == loading_status:
                    specificity += 2
                if row[11] and str(row[11]) == transport_form:
                    specificity += 2

                # Date range validation
                try:
                    service_date_int = int(service_date.replace('-', ''))
                    valid_from = int(row[15]) if row[15] else 0
                    valid_to = int(row[16]) if row[16] else 99991231

                    if not (valid_from <= service_date_int <= valid_to):
                        continue  # No match
                except (ValueError, TypeError):
                    pass

                # If we have any specificity, this is a match
                if specificity > 0 and row[17]:
                    matches.append({
                        'specificity': specificity,
                        'price': float(row[17]),
                        'price_unit': row[18] if row[18] else 'Container',
                        'offer_number': row[0],
                        'customer_number': row[2],
                        'customer_group': row[1],
                        'matched_criteria': specificity
                    })

            # Sort by specificity (highest first)
            matches.sort(key=lambda x: x['specificity'], reverse=True)

            if matches:
                best_match = matches[0]
                logger.info(f"Main service price match (specificity={best_match['specificity']}): €{best_match['price']}")
                return best_match

            logger.warning(f"No main service price match for: {customer_code}, {weight_class}, {direction}")
            return None

        except Exception as e:
            logger.error(f"Error getting main service price: {e}", exc_info=True)
            return None

    def get_main_service_price(self, offer_code: str, weight_class: str,
                               direction: str, **kwargs) -> Optional[float]:
        """
        Get main service price based on criteria

        Args:
            offer_code: Offer/customer code
            weight_class: Weight classification (20A, 20B, 40A, etc.)
            direction: Transport direction (Export, Import, etc.)

        Returns:
            Price in EUR or None if not found
        """

        # Load main service prices
        price_data = self.load_price_file("main_service_prices.xlsx")
        if not price_data:
            logger.warning("Main service prices not loaded")
            return None

        # Find matching price
        for price_entry in price_data['prices']:
            data = price_entry['data']

            # Match criteria
            offer_match = str(data.get('Angebotsnummer')) == str(offer_code) or data.get('Angebotsnummer') == 'alle'
            weight_match = data.get('Gewichts-klasse') == weight_class or data.get('Gewichtsklasse') == weight_class
            direction_match = data.get('Richtung') == direction

            if offer_match and weight_match and direction_match:
                price = data.get('Preis')
                if price is not None:
                    logger.info(f"Main service price: {weight_class} {direction} → €{price}")
                    return float(price)

        # No match found
        logger.warning(f"No main service price found for: {offer_code}, {weight_class}, {direction}")
        return None

    def get_additional_service_price_advanced(self, service_code: str, customer_code: str = None,
                                             customer_group: str = None, departure_station: str = None,
                                             destination_station: str = None, loading_status: str = None,
                                             transport_form: str = None, container_length: str = None,
                                             service_date: str = None, quantity: int = 1,
                                             **kwargs) -> Optional[Dict[str, Any]]:
        """
        Get additional service price with advanced matching and per-unit pricing support

        Args:
            service_code: Service code (e.g., "123", "456", "789")
            All matching criteria from pricing XLSX
            quantity: Number of units (default 1)

        Returns:
            Dictionary with price details:
            {
                'price_per_unit': float,
                'quantity': int,
                'total_price': float,
                'price_basis': 'Container' or 'Einheit',
                'specificity': int
            }
        """

        # Load from shared folder - fix path resolution
        rating_service_dir = self.prices_dir.parent if 'price-tables' in str(self.prices_dir) else self.prices_dir
        project_root = rating_service_dir.parent.parent  # services/rating -> services -> project_root
        shared_path = project_root / "shared" / "rules" / "6_Preistabelle_Nebenleistungen.xlsx"

        if not shared_path.exists():
            logger.warning(f"Additional service price table not found at {shared_path}")
            return None

        try:
            import openpyxl
            wb = openpyxl.load_workbook(shared_path)
            ws = wb.active

            # Expected columns:
            # 0: NGB-Code, 1: NGB-Name (DE), 2: Kundennummer, 3: Kundengruppe,
            # 4: Angebotsnummer, 5: Land Bahnstelle Versand, 6: Bahnstellennummer Versand,
            # 7: Land Bahnstelle Empfang, 8: Bahnstellennummer Empfang,
            # 9: Ladezustand, 10: Verkehrsform, ...
            # 15: Container Länge, 16: gültig von, 17: gültig bis, 18: Preisbasis, 19: Preis

            matches = []

            for row_idx in range(2, ws.max_row + 1):
                row = []
                for col in range(1, ws.max_column + 1):
                    row.append(ws.cell(row_idx, col).value)

                if not any(row):
                    continue

                # Check service code match (required)
                if not row[0] or str(row[0]) != str(service_code):
                    continue

                # Calculate specificity score
                specificity = 0

                # Customer number match = +1000
                if customer_code and row[2] and str(row[2]) == str(customer_code):
                    specificity += 1000
                # Customer group match = +100
                elif customer_group and row[3] and str(row[3]) == str(customer_group):
                    specificity += 100

                # Station matches = +10 each
                if departure_station and row[6] and str(row[6]) == str(departure_station):
                    specificity += 10
                if destination_station and row[8] and str(row[8]) == str(destination_station):
                    specificity += 10

                # Optional matches
                if loading_status and row[9] and str(row[9]) == loading_status:
                    specificity += 2
                if transport_form and row[10] and str(row[10]) == transport_form:
                    specificity += 2
                if container_length and row[15] and str(row[15]) == str(container_length):
                    specificity += 2

                # Date range validation
                if service_date and row[16] and row[17]:
                    try:
                        service_date_int = int(service_date.replace('-', ''))
                        valid_from = int(row[16]) if row[16] else 0
                        valid_to = int(row[17]) if row[17] else 99991231

                        if not (valid_from <= service_date_int <= valid_to):
                            continue  # No match
                    except (ValueError, TypeError):
                        pass

                # Extract price and basis
                if row[19]:
                    price_per_unit = float(row[19])
                    price_basis = str(row[18]) if row[18] else 'Container'

                    # Calculate total price
                    if price_basis.lower() in ['einheit', 'unit', 'per_unit']:
                        total_price = price_per_unit * quantity
                    else:
                        total_price = price_per_unit  # Container-based, quantity doesn't apply

                    matches.append({
                        'specificity': specificity,
                        'price_per_unit': price_per_unit,
                        'quantity': quantity if price_basis.lower() in ['einheit', 'unit', 'per_unit'] else 1,
                        'total_price': total_price,
                        'price_basis': price_basis,
                        'service_name': row[1] if row[1] else f"Service {service_code}"
                    })

            # Sort by specificity (highest first)
            matches.sort(key=lambda x: x['specificity'], reverse=True)

            if matches:
                best_match = matches[0]
                logger.info(f"Additional service {service_code} price match: €{best_match['price_per_unit']} × {best_match['quantity']} = €{best_match['total_price']}")
                return best_match

            logger.warning(f"No additional service price match for: {service_code}")
            return None

        except Exception as e:
            logger.error(f"Error getting additional service price: {e}", exc_info=True)
            return None

    def get_additional_service_price(self, service_code: str, container_length: str = None,
                                     **kwargs) -> Optional[float]:
        """
        Get additional service price

        Args:
            service_code: Service code (e.g., 123, 456)
            container_length: Container length for size-dependent pricing (20, 40)

        Returns:
            Price in EUR or None if not found
        """

        # Load additional service prices
        price_data = self.load_price_file("additional_service_prices.xlsx")
        if not price_data:
            logger.warning("Additional service prices not loaded")
            return None

        # Find matching price
        for price_entry in price_data['prices']:
            data = price_entry['data']

            # Match service code
            code_match = str(data.get('Code')) == str(service_code)

            if code_match:
                # Check container length if specified
                if container_length:
                    price_container_length = data.get('Container Länge') or data.get('Container Length')
                    if price_container_length and str(price_container_length) != str(container_length):
                        continue

                price = data.get('Preis')
                if price is not None:
                    logger.info(f"Additional service price: {service_code} ({container_length}ft) → €{price}")
                    return float(price)

        # No match found
        logger.warning(f"No additional service price found for: {service_code}, {container_length}ft")
        return None

    def get_all_prices(self, file_name: str) -> List[Dict]:
        """Get all prices from a file"""
        price_data = self.load_price_file(file_name)
        if price_data:
            return price_data['prices']
        return []

    def reload_prices(self, force: bool = True) -> Dict[str, bool]:
        """Reload all cached prices, optionally forcing reload regardless of modification time"""
        results = {}
        cached_files = list(self._price_cache.keys())

        if force:
            # Clear all caches for force reload
            self._price_cache.clear()
            self._file_mtimes.clear()
            logger.info(f"Force reloading {len(cached_files)} price files")

        # Reload each file (will auto-detect modifications if not forced)
        for file_name in cached_files:
            price_data = self.load_price_file(file_name, force_reload=force)
            results[file_name] = price_data is not None

        return results

    def get_price_info(self, file_name: str) -> Optional[Dict[str, Any]]:
        """Get information about a price file"""
        file_path = self.prices_dir / file_name

        if not file_path.exists():
            return None

        price_data = self.load_price_file(file_name)

        return {
            'file_name': file_name,
            'path': str(file_path),
            'exists': True,
            'loaded': price_data is not None,
            'sheets': list(price_data['sheets'].keys()) if price_data else [],
            'prices_count': len(price_data['prices']) if price_data else 0,
            'size': file_path.stat().st_size,
            'modified': file_path.stat().st_mtime
        }