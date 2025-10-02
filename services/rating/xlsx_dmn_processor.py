#!/usr/bin/env python3
"""
Dynamic XLSX DMN Rule Processor
Direct processor for XLSX rule files from requirement documents
Alternative to pyDMNrules that works with the specific format of requirement docs
"""

import openpyxl
from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
import logging
import re

logger = logging.getLogger(__name__)

class XLSXDMNProcessor:
    """
    Direct XLSX DMN rule processor that can work with requirement document formats
    """

    def __init__(self, rules_dir: Path):
        self.rules_dir = Path(rules_dir)
        self._rule_cache: Dict[str, Dict] = {}
        self._file_mtimes: Dict[str, float] = {}  # Track file modification times

    def load_rule_file(self, file_name: str, force_reload: bool = False) -> Optional[Dict]:
        """Load and parse an XLSX rule file with automatic modification detection"""

        file_path = self.rules_dir / file_name

        if not file_path.exists():
            logger.warning(f"Rule file not found: {file_path}")
            return None

        # Check if file was modified since last load
        try:
            current_mtime = file_path.stat().st_mtime
        except Exception as e:
            logger.error(f"Failed to get file modification time for {file_name}: {e}")
            current_mtime = 0

        # Use cache if file hasn't been modified and not forcing reload
        if not force_reload and file_name in self._rule_cache:
            cached_mtime = self._file_mtimes.get(file_name, 0)
            if current_mtime <= cached_mtime:
                logger.debug(f"Using cached rules for {file_name}")
                return self._rule_cache[file_name]
            else:
                logger.info(f"File {file_name} modified (cached: {cached_mtime}, current: {current_mtime}), reloading")

        # Load file and cache with modification time
        try:
            wb = openpyxl.load_workbook(file_path)
            rule_data = self._parse_workbook(wb, file_name)
            self._rule_cache[file_name] = rule_data
            self._file_mtimes[file_name] = current_mtime
            logger.info(f"Loaded rules from {file_name} (mtime: {current_mtime})")
            return rule_data

        except Exception as e:
            logger.error(f"Failed to load rule file {file_name}: {e}")
            return None

    def _parse_workbook(self, wb: openpyxl.Workbook, file_name: str) -> Dict:
        """Parse a workbook and extract rule data"""

        rule_data = {
            'file_name': file_name,
            'sheets': {},
            'rules': []
        }

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = self._parse_sheet(sheet, sheet_name)
            rule_data['sheets'][sheet_name] = sheet_data

            # Extract rules from sheet
            if sheet_data['headers'] and sheet_data['rows']:
                rules = self._extract_rules_from_sheet(sheet_data, sheet_name)
                rule_data['rules'].extend(rules)

        return rule_data

    def _parse_sheet(self, sheet, sheet_name: str) -> Dict:
        """
        Parse a worksheet and extract structured data

        For DMN decision tables, the format is:
        - Row 1: Table name in A1
        - Row 2: Hit policy in A2, headers in B2, C2, etc.
        - Row 3+: Empty in col A, data in B3+, C3+, etc.

        For regular sheets (Glossary, Decision):
        - Row 1: Headers
        - Row 2+: Data
        """

        sheet_data = {
            'name': sheet_name,
            'headers': [],
            'rows': [],
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'is_dmn_table': False
        }

        if sheet.max_row < 1:
            return sheet_data

        # Check if this is a DMN decision table
        # DMN tables have a table name in A1 and hit policy in A2 (row 2, col 1)
        cell_a1 = sheet.cell(1, 1).value
        cell_a2 = sheet.cell(2, 1).value

        if cell_a2 and str(cell_a2).strip() in ['U', 'A', 'P', 'F', 'R', 'O', 'C', 'C+', 'C<', 'C>', 'C#']:
            # This is a DMN decision table
            sheet_data['is_dmn_table'] = True
            sheet_data['table_name'] = str(cell_a1).strip() if cell_a1 else sheet_name
            sheet_data['hit_policy'] = str(cell_a2).strip()

            # Get headers from row 2, starting from column 2
            headers = []
            for col in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(2, col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Col_{col}")

            sheet_data['headers'] = headers

            # Get data rows starting from row 3, column 2
            for row in range(3, sheet.max_row + 1):
                row_data = []
                for col in range(2, sheet.max_column + 1):
                    cell_value = sheet.cell(row, col).value
                    if cell_value is not None:
                        # Clean quoted strings
                        value_str = str(cell_value).strip()
                        # Remove outer quotes if present
                        if value_str.startswith('"') and value_str.endswith('"'):
                            value_str = value_str[1:-1]
                        row_data.append(value_str)
                    else:
                        row_data.append("")

                # Skip empty rows
                if any(cell for cell in row_data if cell):
                    sheet_data['rows'].append(row_data)
        else:
            # Regular sheet format (Glossary, Decision, etc.)
            # Get headers from first row
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(1, col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Col_{col}")

            sheet_data['headers'] = headers

            # Get data rows
            for row in range(2, sheet.max_row + 1):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row, col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value).strip())
                    else:
                        row_data.append("")

                # Skip empty rows
                if any(cell for cell in row_data if cell):
                    sheet_data['rows'].append(row_data)

        return sheet_data

    def _extract_rules_from_sheet(self, sheet_data: Dict, sheet_name: str) -> List[Dict]:
        """Extract rules from sheet data based on known patterns"""

        rules = []
        headers = sheet_data['headers']
        rows = sheet_data['rows']

        for i, row in enumerate(rows):
            rule = {
                'sheet': sheet_name,
                'rule_id': f"{sheet_name}_{i+1}",
                'conditions': {},
                'outputs': {},
                'raw_row': row
            }

            # Map row data to headers
            for j, header in enumerate(headers):
                if j < len(row):
                    value = row[j]
                    if value and value != '':
                        # Clean up quoted values
                        cleaned_value = value.strip('"')
                        rule['conditions'][header] = cleaned_value
                        rule['outputs'][header] = cleaned_value

            rules.append(rule)

        return rules

    def evaluate_trip_type(self, trucking_code: str, **kwargs) -> Optional[str]:
        """Evaluate trip type based on trucking code"""

        # Load trip type rules
        rule_data = self.load_rule_file("trip_type.dmn.xlsx")
        if not rule_data:
            logger.warning("Trip type rules not loaded")
            return None

        # Find matching rule
        for rule in rule_data['rules']:
            # Check different possible column names for trucking code
            trucking_code_fields = ['Trucking Code', 'TruckingCode', 'Trucking_Code']
            trip_type_fields = ['TypeOfTrip', 'Trip Type', 'Type Of Trip', 'Fahrttyp']

            for tc_field in trucking_code_fields:
                if tc_field in rule['conditions']:
                    rule_code = rule['conditions'][tc_field]
                    if rule_code == trucking_code:
                        # Return the trip type
                        for tt_field in trip_type_fields:
                            if tt_field in rule['outputs']:
                                result = rule['outputs'][tt_field]
                                logger.debug(f"Trip type: {trucking_code} -> {result}")
                                return result

        # Default fallback
        default_trip_type = "Zustellung"
        logger.debug(f"Trip type fallback: {trucking_code} -> {default_trip_type}")
        return default_trip_type

    def evaluate_weight_class(self, container_length: str, gross_weight: float, preisraster: str = "N", **kwargs) -> Optional[str]:
        """
        Evaluate weight class based on container length and gross weight

        Args:
            container_length: Container length ("20" or "40")
            gross_weight: Gross weight in kg
            preisraster: Price grid identifier (default "N")

        Returns:
            Weight classification (e.g., "20A", "20B", "40A", etc.)
        """

        # Load weight classification rules
        rule_data = self.load_rule_file("5_Regeln_Gewichtsklassen.xlsx")
        if not rule_data:
            logger.warning("Weight classification rules not loaded")
            return None

        # Convert weight from kg to tons for comparison
        gross_weight_tons = gross_weight / 1000.0

        # Find matching rule
        for rule in rule_data['rules']:
            # Check different possible column names
            preisraster_fields = ['Preisraster', 'Price Grid', 'Grid']
            length_fields = ['Länge', 'Laenge', 'Length', 'Container Length']
            weight_fields = ['Gewicht', 'Weight', 'GrossWeight', 'Gross Weight']
            class_fields = ['WeightClass', 'Weight Class', 'Gewichtsklasse', 'Classification']

            preisraster_match = True  # Default to true if not specified
            length_match = False
            weight_match = False

            # Check preisraster (if specified in rule)
            for pf in preisraster_fields:
                if pf in rule['conditions']:
                    rule_preisraster = rule['conditions'][pf].strip('"\'')
                    if rule_preisraster != '-' and rule_preisraster != '':
                        preisraster_match = (rule_preisraster == preisraster)
                    break

            # Check length condition
            for lf in length_fields:
                if lf in rule['conditions']:
                    rule_length = rule['conditions'][lf].strip('"\'')
                    if rule_length == container_length or rule_length == '-':
                        length_match = True
                        break

            # Check weight condition using FEEL expression evaluation
            for wf in weight_fields:
                if wf in rule['conditions']:
                    rule_weight = rule['conditions'][wf]
                    if self._evaluate_weight_condition(gross_weight_tons, rule_weight):
                        weight_match = True
                        break

            # If all conditions match, return the classification
            if preisraster_match and length_match and weight_match:
                for cf in class_fields:
                    if cf in rule['outputs']:
                        result = rule['outputs'][cf]
                        logger.info(f"Weight class: {container_length}ft, {gross_weight}kg ({gross_weight_tons}t) -> {result}")
                        return result

        # No matching rule found
        logger.warning(f"No matching weight class rule for: {container_length}ft, {gross_weight}kg")
        return None

    def _evaluate_weight_condition(self, actual_weight: float, rule_condition: str) -> bool:
        """
        Evaluate if an actual weight value matches a FEEL rule condition

        Args:
            actual_weight: The actual weight value (in tons)
            rule_condition: FEEL expression like "<=20", ">20", "[10..20]"

        Returns:
            True if the weight matches the condition
        """

        if rule_condition == '-' or rule_condition == '':
            return True

        rule_condition = rule_condition.strip()

        try:
            # Handle range expressions like "[10..20]", "]10..20]", "[10..20[", etc.
            if '..' in rule_condition:
                # Extract bracket types and values
                left_inclusive = rule_condition[0] == '['
                right_inclusive = rule_condition[-1] == ']'

                # Remove brackets and extract range values
                range_str = rule_condition.strip('[]')
                parts = range_str.split('..')
                if len(parts) != 2:
                    logger.warning(f"Invalid range expression: {rule_condition}")
                    return False

                lower = float(parts[0].strip())
                upper = float(parts[1].strip())

                # Check if weight is in range
                if left_inclusive and right_inclusive:
                    return lower <= actual_weight <= upper
                elif left_inclusive and not right_inclusive:
                    return lower <= actual_weight < upper
                elif not left_inclusive and right_inclusive:
                    return lower < actual_weight <= upper
                else:
                    return lower < actual_weight < upper

            # Handle comparison expressions
            elif rule_condition.startswith('<='):
                threshold = float(rule_condition[2:].strip())
                return actual_weight <= threshold

            elif rule_condition.startswith('>='):
                threshold = float(rule_condition[2:].strip())
                return actual_weight >= threshold

            elif rule_condition.startswith('<'):
                threshold = float(rule_condition[1:].strip())
                return actual_weight < threshold

            elif rule_condition.startswith('>'):
                threshold = float(rule_condition[1:].strip())
                return actual_weight > threshold

            elif rule_condition.startswith('='):
                threshold = float(rule_condition[1:].strip())
                return actual_weight == threshold

            # Try direct numeric comparison
            else:
                try:
                    threshold = float(rule_condition)
                    return actual_weight == threshold
                except ValueError:
                    logger.warning(f"Could not parse weight condition: {rule_condition}")
                    return False

        except (ValueError, IndexError) as e:
            logger.warning(f"Error evaluating weight condition '{rule_condition}' for weight {actual_weight}: {e}")
            return False

        return False

    def evaluate_service_determination(self, verkehrsform: str, gefahrgut: bool,
                                      gueltig_von: str = None, gueltig_bis: str = None,
                                      **kwargs) -> Optional[List[int]]:
        """Evaluate which additional services apply (legacy method for backward compatibility)"""
        # Load service determination rules
        rule_data = self.load_rule_file("service_determination.dmn.xlsx")
        if not rule_data:
            logger.warning("Service determination rules not loaded")
            return None

        applicable_services = []

        # Find matching rules
        for rule in rule_data['rules']:
            # Check conditions
            verkehrsform_fields = ['Verkehrsform', 'Transport Type', 'TransportType']
            gefahrgut_fields = ['Gefahrgut', 'Dangerous Goods', 'DangerousGoods']
            service_fields = ['AdditionalServiceCode', 'Service Code', 'ServiceCode']

            verkehrsform_match = True  # Default to true
            gefahrgut_match = True     # Default to true

            # Check Verkehrsform
            for vf in verkehrsform_fields:
                if vf in rule['conditions']:
                    rule_verkehrsform = rule['conditions'][vf]
                    if rule_verkehrsform != '-' and rule_verkehrsform != '':
                        verkehrsform_match = (rule_verkehrsform == verkehrsform)
                    break

            # Check Gefahrgut
            for gf in gefahrgut_fields:
                if gf in rule['conditions']:
                    rule_gefahrgut = rule['conditions'][gf]
                    if rule_gefahrgut != '-' and rule_gefahrgut != '':
                        # Handle boolean conversion
                        rule_gefahrgut_bool = rule_gefahrgut.lower() in ['true', '1', 'yes']
                        gefahrgut_match = (rule_gefahrgut_bool == gefahrgut)
                    break

            # If conditions match, extract service code
            if verkehrsform_match and gefahrgut_match:
                for sf in service_fields:
                    if sf in rule['outputs']:
                        service_code = rule['outputs'][sf]
                        try:
                            service_int = int(service_code)
                            if service_int not in applicable_services:
                                applicable_services.append(service_int)
                        except ValueError:
                            pass

        logger.debug(f"Service determination: {verkehrsform}, {gefahrgut} -> {applicable_services}")
        return applicable_services

    def evaluate_service_determination_full(self, order_context: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Evaluate service determination with full COLLECT policy support
        Reads from 4_Regeln_Leistungsermittlung.xlsx in shared/rules/

        Args:
            order_context: Dictionary with order details:
                - service_type: "Hauptleistung Transport", etc.
                - loading_status: "beladen" or "leer"
                - transport_type: "KV", "KVS", etc.
                - dangerous_goods: boolean
                - customs_procedure: "N1", "T1-NCTS", etc.
                - departure_country: "DE", etc.
                - departure_station: railway station number
                - destination_country: "DE", "US", etc.
                - destination_station: railway station number
                - service_date: date in YYYYMMDD format

        Returns:
            List of dictionaries with service details:
            [
                {"code": 111, "name": "Zuschlag 1", "rule_matched": "Rule 1"},
                {"code": 222, "name": "Zuschlag 2", "rule_matched": "Rule 2"},
                ...
            ]
        """

        # Load service determination rules from configured rules directory
        # The rules_dir should point to dmn-rules symlink which points to shared/rules/
        rules_path = self.rules_dir / "4_Regeln_Leistungsermittlung.xlsx"

        if not rules_path.exists():
            logger.warning(f"Service determination rules not found at {rules_path}")
            logger.debug(f"Rules directory: {self.rules_dir}, exists: {self.rules_dir.exists()}")
            return []

        try:
            wb = openpyxl.load_workbook(rules_path)
            ws = wb.active

            # Parse the XLSX structure
            # Row 1: Headers
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(1, col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Col_{col}")

            # Expected columns based on analysis:
            # 0: Leistung
            # 1: Ladezustand
            # 2: Verkehrsform
            # 3: Gefahrgut vorhanden
            # 4: Zollverfahren
            # 5: Land Bahnstelle Versand
            # 6: Bahnstellennummer Versand
            # 7: Land Bahnstelle Empfang
            # 8: Bahnstellenummer Empfang
            # 9: gültig von
            # 10: gültig bis
            # 11: NGB-Code
            # 12: NGB-Name (DE)

            matched_services = []

            # Iterate through rules (row 2 onwards)
            for row_idx in range(2, ws.max_row + 1):
                row_values = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row_idx, col).value
                    row_values.append(cell_value)

                # Skip empty rows
                if not any(row_values):
                    continue

                # Extract rule conditions
                rule_service_type = row_values[0] if len(row_values) > 0 else None
                rule_loading = row_values[1] if len(row_values) > 1 else None
                rule_transport = row_values[2] if len(row_values) > 2 else None
                rule_dangerous = row_values[3] if len(row_values) > 3 else None
                rule_customs = row_values[4] if len(row_values) > 4 else None
                rule_dep_country = row_values[5] if len(row_values) > 5 else None
                rule_dep_station = row_values[6] if len(row_values) > 6 else None
                rule_dest_country = row_values[7] if len(row_values) > 7 else None
                rule_dest_station = row_values[8] if len(row_values) > 8 else None
                rule_valid_from = row_values[9] if len(row_values) > 9 else None
                rule_valid_to = row_values[10] if len(row_values) > 10 else None
                rule_ngb_code = row_values[11] if len(row_values) > 11 else None
                rule_ngb_name = row_values[12] if len(row_values) > 12 else None

                # Match logic: None or empty = wildcard (matches anything)
                matches = True

                # Service type match
                if rule_service_type and str(rule_service_type).strip().strip('"'):
                    order_service_type = order_context.get('service_type', '')
                    if str(rule_service_type).strip('"') != order_service_type:
                        matches = False

                # Loading status match
                if matches and rule_loading and str(rule_loading).strip().strip('"'):
                    order_loading = order_context.get('loading_status', '')
                    if str(rule_loading).strip('"') != order_loading:
                        matches = False

                # Transport type match
                if matches and rule_transport and str(rule_transport).strip().strip('"'):
                    order_transport = order_context.get('transport_type', '')
                    rule_transport_clean = str(rule_transport).strip('"')
                    # KVS matches both KV and KVS
                    if rule_transport_clean not in [order_transport, 'KV'] and order_transport != rule_transport_clean:
                        matches = False

                # Dangerous goods match
                if matches and rule_dangerous and str(rule_dangerous).lower() == 'true':
                    if not order_context.get('dangerous_goods', False):
                        matches = False

                # Customs procedure match
                if matches and rule_customs and str(rule_customs).strip().strip('"'):
                    order_customs = order_context.get('customs_procedure', '')
                    if str(rule_customs).strip('"') != order_customs:
                        matches = False

                # Departure country match
                if matches and rule_dep_country and str(rule_dep_country).strip().strip('"'):
                    order_dep_country = order_context.get('departure_country', '')
                    if str(rule_dep_country).strip('"') != order_dep_country:
                        matches = False

                # Departure station match
                if matches and rule_dep_station and str(rule_dep_station).strip().strip('"'):
                    order_dep_station = order_context.get('departure_station', '')
                    if str(rule_dep_station).strip('"') != order_dep_station:
                        matches = False

                # Destination country match
                if matches and rule_dest_country and str(rule_dest_country).strip().strip('"'):
                    order_dest_country = order_context.get('destination_country', '')
                    if str(rule_dest_country).strip('"') != order_dest_country:
                        matches = False

                # Destination station match
                if matches and rule_dest_station and str(rule_dest_station).strip().strip('"'):
                    order_dest_station = order_context.get('destination_station', '')
                    if str(rule_dest_station).strip('"') != order_dest_station:
                        matches = False

                # Date range validation
                if matches and rule_valid_from and rule_valid_to:
                    order_date = order_context.get('service_date')
                    if order_date:
                        try:
                            order_date_int = int(order_date)
                            valid_from_int = int(rule_valid_from) if rule_valid_from else 0
                            valid_to_int = int(rule_valid_to) if rule_valid_to else 99991231

                            if not (valid_from_int <= order_date_int <= valid_to_int):
                                matches = False
                        except (ValueError, TypeError):
                            pass

                # If all conditions match, add the service
                if matches and rule_ngb_code:
                    service_dict = {
                        'code': int(rule_ngb_code),
                        'name': str(rule_ngb_name) if rule_ngb_name else f"Service {rule_ngb_code}",
                        'rule_matched': f"Row {row_idx}"
                    }
                    matched_services.append(service_dict)
                    logger.debug(f"Service determination match: Row {row_idx} -> Service {rule_ngb_code}: {rule_ngb_name}")

            logger.info(f"Service determination: {len(matched_services)} services matched")
            return matched_services

        except Exception as e:
            logger.error(f"Error evaluating service determination: {e}", exc_info=True)
            return []

    def determine_service_789_from_123(self, services: List[Dict[str, Any]], additional_service_123_data: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """
        Auto-determine service 789 (waiting time) when service 123 (Zustellung Export) is present

        Based on artefact 6_3_Abrechnungsbeleg_Nebenleistung_2.json:
        - Service 789: Wartezeit Export
        - AmountBrutto: 8 units
        - AmountNetto: 5 units (used for pricing)
        - Unit: "Einheit"
        - Price: €50 per unit
        - Total: 5 units × €50 = €250

        Args:
            services: List of service dictionaries from evaluate_service_determination_full
            additional_service_123_data: Optional data about service 123 with quantity info

        Returns:
            Updated services list with service 789 added if service 123 present
        """

        # Check if service 123 is in the list
        has_service_123 = any(s.get('code') == 123 for s in services)

        if not has_service_123:
            logger.debug("Service 123 not found, skipping service 789 auto-determination")
            return services

        # Check if service 789 already exists (avoid duplicates)
        has_service_789 = any(s.get('code') == 789 for s in services)

        if has_service_789:
            logger.debug("Service 789 already exists, skipping auto-determination")
            return services

        # Determine quantity for service 789
        # From artefact: AmountNetto = 5 units
        quantity_netto = 5
        quantity_brutto = 8

        if additional_service_123_data:
            # If we have data about service 123, use its quantity info
            quantity_netto = additional_service_123_data.get('amount_netto', 5)
            quantity_brutto = additional_service_123_data.get('amount_brutto', 8)

        # Add service 789
        service_789 = {
            'code': 789,
            'name': 'Wartezeit Export',
            'rule_matched': 'Auto-determined from service 123',
            'quantity_netto': quantity_netto,
            'quantity_brutto': quantity_brutto,
            'unit': 'Einheit',
            'price_per_unit': 50.0,  # From hardcoded_prices_383.sql
            'total_amount': quantity_netto * 50.0  # €250
        }

        services.append(service_789)
        logger.info(f"Auto-determined service 789 (Wartezeit Export): {quantity_netto} units × €50 = €{service_789['total_amount']}")

        return services

    def get_available_rules(self) -> List[str]:
        """Get list of available rule files"""
        if not self.rules_dir.exists():
            return []

        xlsx_files = []
        for file_path in self.rules_dir.glob("*.xlsx"):
            xlsx_files.append(file_path.name)

        return sorted(xlsx_files)

    def reload_rules(self, force: bool = True) -> Dict[str, bool]:
        """Reload all cached rules, optionally forcing reload regardless of modification time"""
        results = {}
        cached_files = list(self._rule_cache.keys())

        if force:
            # Clear all caches for force reload
            self._rule_cache.clear()
            self._file_mtimes.clear()
            logger.info(f"Force reloading {len(cached_files)} rule files")

        # Reload each file (will auto-detect modifications if not forced)
        for file_name in cached_files:
            rule_data = self.load_rule_file(file_name, force_reload=force)
            results[file_name] = rule_data is not None

        return results

    def get_rule_info(self, file_name: str) -> Optional[Dict[str, Any]]:
        """Get information about a rule file"""
        file_path = self.rules_dir / file_name

        if not file_path.exists():
            return None

        rule_data = self.load_rule_file(file_name)

        return {
            'file_name': file_name,
            'path': str(file_path),
            'exists': True,
            'loaded': rule_data is not None,
            'sheets': list(rule_data['sheets'].keys()) if rule_data else [],
            'rules_count': len(rule_data['rules']) if rule_data else 0,
            'size': file_path.stat().st_size,
            'modified': file_path.stat().st_mtime
        }